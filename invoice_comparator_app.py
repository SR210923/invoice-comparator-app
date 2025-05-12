import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Invoice Comparator", layout="wide")
st.title("🧾 Purchase Invoice Comparator")

st.write("Upload two invoice Excel files. We'll highlight any differences in price, quantity, or missing items.")

# File uploads
file1 = st.file_uploader("Upload Invoice A", type=["xlsx"])
file2 = st.file_uploader("Upload Invoice B", type=["xlsx"])

if file1 and file2:
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Normalize column names (you can customize this)
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    key_column = "Item Code"
    quantity_column = "Quantity"
    price_column = "Price"

    if key_column in df1.columns and key_column in df2.columns:
        merged = pd.merge(df1, df2, on=key_column, how='outer', suffixes=("_A", "_B"))

        # Identify differences
        merged['Qty Diff'] = merged[f'{quantity_column}_A'] != merged[f'{quantity_column}_B']
        merged['Price Diff'] = merged[f'{price_column}_A'] != merged[f'{price_column}_B']
        merged['Missing In A'] = merged[f'{quantity_column}_A'].isna()
        merged['Missing In B'] = merged[f'{quantity_column}_B'].isna()

        st.subheader("Comparison Table")
        st.dataframe(merged)

        # Create Excel output with highlights
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"

        # Colors
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        blue_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")

        # Write headers
        ws.append(list(merged.columns))

        # Write rows with styling
        for _, row in merged.iterrows():
            ws_row = list(row)
            ws.append(ws_row)
            current_row = ws.max_row
            for i, col_name in enumerate(merged.columns):
                cell = ws.cell(row=current_row, column=i + 1)
                if col_name == 'Price Diff' and row['Price Diff']:
                    cell.fill = red_fill
                elif col_name == 'Qty Diff' and row['Qty Diff']:
                    cell.fill = yellow_fill
                elif (col_name == f'{quantity_column}_A' and row['Missing In A']) or                      (col_name == f'{quantity_column}_B' and row['Missing In B']):
                    cell.fill = blue_fill

        # Prepare download
        output = BytesIO()
        wb.save(output)
        st.download_button("📥 Download Comparison Report", data=output.getvalue(), file_name="Invoice_Comparison.xlsx")

    else:
        st.error(f"'{key_column}' must be present in both files. Please check column names.")
